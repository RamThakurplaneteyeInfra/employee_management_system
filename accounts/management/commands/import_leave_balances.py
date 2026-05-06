"""
Bulk-update casual / earn / menstrual leave balances on `LeaveSummary` from a sheet.

Accepts a CSV (.csv) or Excel (.xlsx / .xlsm) file with these columns
(header row required, case-insensitive, extra columns are ignored):

    username, casual_leaves, earn_leaves, menstrual_leaves

Behaviour
---------
* Match rows by `User.username` (which is also `Profile.Employee_id`).
* Decimal values supported (e.g. 1.5, 4.5).
* Blank / missing cell for a leave column => that field is left unchanged for that user.
* If `LeaveSummary` row does not exist for a user, it is reported and skipped
  (you can pass --create-missing to auto-create with default 0 totals).
* Wrapped in a single transaction; nothing is written if any row fails parsing.
* `--dry-run` prints the diff but does not write to DB.

Untouched columns: total_leaves, used_leaves, emergency_leaves, and every other
table in the project. The leave-application flow (leave_views.py) is not invoked
or affected in any way.

Examples
--------
    python manage.py import_leave_balances --file leave_balances.csv --dry-run
    python manage.py import_leave_balances --file leave_balances.csv
    python manage.py import_leave_balances --file leave_balances.xlsx --create-missing
"""
from __future__ import annotations

import csv
import os
from decimal import Decimal, InvalidOperation
from typing import Iterable, Optional

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import LeaveSummary


LEAVE_COLUMNS = ("casual_leaves", "earn_leaves", "menstrual_leaves")
USERNAME_HEADERS = ("username", "user", "employee_id", "employeeid", "emp_id")


def _norm(s: object) -> str:
    return ("" if s is None else str(s)).strip()


def _parse_decimal(raw: object, column: str, row_no: int) -> Optional[Decimal]:
    """Return Decimal or None for blank cells. Raise CommandError on bad value."""
    text = _norm(raw)
    if text == "":
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        raise CommandError(
            f"Row {row_no}: column '{column}' has invalid number '{text}'."
        )


def _read_rows(file_path: str) -> Iterable[dict]:
    """Yield row dicts (header-keyed) from CSV or XLSX."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                yield {(k or "").strip().lower(): v for k, v in row.items()}
        return
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required to read .xlsx files. Install with `pip install openpyxl`."
            ) from exc
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return
        norm_header = [(_norm(c) or "").lower() for c in header]
        for raw in rows_iter:
            if raw is None:
                continue
            yield {norm_header[i]: raw[i] if i < len(raw) else None for i in range(len(norm_header))}
        return
    raise CommandError(f"Unsupported file extension '{ext}'. Use .csv or .xlsx.")


def _resolve_username_key(headers: Iterable[str]) -> Optional[str]:
    lowered = {h.lower() for h in headers}
    for candidate in USERNAME_HEADERS:
        if candidate in lowered:
            return candidate
    return None


class Command(BaseCommand):
    help = (
        "Bulk-update casual_leaves / earn_leaves / menstrual_leaves on LeaveSummary "
        "from a CSV or XLSX file (header row: username, casual_leaves, earn_leaves, menstrual_leaves)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Path to .csv or .xlsx file to import.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Show what would change without writing to DB.",
        )
        parser.add_argument(
            "--create-missing",
            action="store_true",
            help="Auto-create LeaveSummary rows for users that do not have one.",
        )

    def handle(self, *args, **options):
        path = options["file"]
        dry_run = bool(options.get("dry_run"))
        create_missing = bool(options.get("create_missing"))

        if not os.path.isfile(path):
            raise CommandError(f"File not found: {path}")

        rows = list(_read_rows(path))
        if not rows:
            self.stdout.write(self.style.WARNING("Sheet is empty; nothing to do."))
            return

        username_key = _resolve_username_key(rows[0].keys())
        if not username_key:
            raise CommandError(
                "Header must contain a username column. "
                f"Accepted: {', '.join(USERNAME_HEADERS)}."
            )

        # ---- Parse + validate every row first; only write at the end ----
        plan: list[tuple[str, dict]] = []  # (username, {field: Decimal})
        skipped_no_user: list[str] = []
        skipped_no_summary: list[str] = []

        users_by_name = {
            u.username: u for u in User.objects.filter(
                username__in=[_norm(r.get(username_key)) for r in rows if _norm(r.get(username_key))]
            )
        }

        for idx, row in enumerate(rows, start=2):  # idx 2 = first data row in spreadsheet
            username = _norm(row.get(username_key))
            if not username:
                self.stdout.write(self.style.WARNING(f"Row {idx}: blank username, skipped."))
                continue

            updates: dict = {}
            for col in LEAVE_COLUMNS:
                value = _parse_decimal(row.get(col), col, idx)
                if value is not None:
                    updates[col] = value

            if not updates:
                self.stdout.write(
                    self.style.NOTICE(
                        f"Row {idx} ({username}): all leave cells blank, nothing to update."
                    )
                )
                continue

            user = users_by_name.get(username)
            if not user:
                skipped_no_user.append(username)
                continue

            plan.append((username, updates))

        if not plan and not skipped_no_user:
            self.stdout.write(self.style.WARNING("No applicable rows; nothing to do."))
            return

        # ---- Pre-fetch existing summaries ----
        existing = {
            ls.user_id: ls
            for ls in LeaveSummary.objects.filter(
                user__username__in=[u for u, _ in plan]
            )
        }
        # `user_id` stores username (because OneToOneField uses to_field="username").

        will_update: list[tuple[LeaveSummary, dict]] = []
        will_create: list[tuple[str, dict]] = []
        for username, updates in plan:
            ls = existing.get(username)
            if ls is None:
                if create_missing:
                    will_create.append((username, updates))
                else:
                    skipped_no_summary.append(username)
                continue
            will_update.append((ls, updates))

        # ---- Report ----
        self.stdout.write(self.style.MIGRATE_HEADING("Import plan"))
        self.stdout.write(f"  to update     : {len(will_update)}")
        self.stdout.write(f"  to create     : {len(will_create)}")
        self.stdout.write(f"  unknown user  : {len(skipped_no_user)}")
        self.stdout.write(f"  no summary    : {len(skipped_no_summary)} (use --create-missing to seed)")

        if skipped_no_user:
            self.stdout.write(self.style.WARNING(
                "  Unknown usernames: " + ", ".join(skipped_no_user)
            ))
        if skipped_no_summary:
            self.stdout.write(self.style.WARNING(
                "  Users without LeaveSummary: " + ", ".join(skipped_no_summary)
            ))

        for ls, upd in will_update:
            diff = ", ".join(
                f"{k}: {getattr(ls, k)} -> {v}" for k, v in upd.items()
            )
            self.stdout.write(f"  UPDATE {ls.user_id}: {diff}")
        for username, upd in will_create:
            diff = ", ".join(f"{k}={v}" for k, v in upd.items())
            self.stdout.write(f"  CREATE {username}: {diff}")

        if dry_run:
            self.stdout.write(self.style.NOTICE("[dry-run] No changes written."))
            return

        # ---- Apply atomically ----
        with transaction.atomic():
            for ls, upd in will_update:
                for k, v in upd.items():
                    setattr(ls, k, v)
                ls.save(update_fields=list(upd.keys()))
            for username, upd in will_create:
                user = users_by_name[username]
                LeaveSummary.objects.create(
                    user=user,
                    total_leaves=0,
                    used_leaves=0,
                    **upd,
                )

        self.stdout.write(self.style.SUCCESS(
            f"Done. Updated {len(will_update)}, created {len(will_create)}."
        ))
