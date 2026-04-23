"""Parse .xlsx rows into structure-entry payloads (first sheet, header row)."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .models import AbstractStructureEntry

# (model_field, aliases...) — header match is case-insensitive, whitespace-collapsed
FIELD_SPECS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("project_name", ("project", "project name", "project_name")),
    ("team_lead_name", ("team lead name", "team_lead_name", "team lead")),
    ("branch", ("branch",)),
    ("date_of_entry", ("date of entry", "date_of_entry", "date")),
    ("route_corridor", ("route / corridor", "route/corridor", "route corridor", "route_corridor", "route")),
    ("sr_no", ("sr. no", "sr no", "sr_no", "serial no", "serial")),
    ("chainage", ("chainage",)),
    ("structure_type", ("structure type", "structure_type", "type", "bridge / structure type")),
    ("length_of_structure", ("length of structure", "length_of_structure", "length")),
    ("span_arrangement", ("span arrangement", "span_arrangement", "span")),
    ("equipment_notes", ("equipment notes", "equipment_notes", "accessible to / equipment notes", "notes")),
    ("inspection_status", ("inspection status", "inspection_status", "inspection")),
    ("remark", ("remark", "remarks")),
    ("las_file_submitted", ("las file submitted", "las_file_submitted", "las submitted")),
    ("reports_available_on_bms_for_las", ("reports available on bms for las", "reports_available_on_bms_for_las", "bms las")),
    ("sar_files_submitted", ("sar files submitted", "sar_files_submitted", "sar submitted")),
    ("reports_available_on_bms_for_sar", ("reports available on bms for sar", "reports_available_on_bms_for_sar", "bms sar")),
)


def _norm_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _header_to_field_map() -> dict[str, str]:
    m: dict[str, str] = {}
    for field, aliases in FIELD_SPECS:
        m[_norm_header(field.replace("_", " "))] = field
        for a in aliases:
            m[_norm_header(a)] = field
    return m


HEADER_TO_FIELD = _header_to_field_map()

# Full key set for DRF so bulk rows never omit keys (avoids "field required" on partial sheets).
ENTRY_PAYLOAD_DEFAULTS: dict[str, Any] = {
    "project_name": "",
    "team_lead_name": "",
    "branch": "INFRA_CORE",
    "date_of_entry": None,
    "route_corridor": "",
    "sr_no": "",
    "chainage": "",
    "structure_type": "",
    "length_of_structure": "",
    "span_arrangement": "",
    "equipment_notes": "",
    "inspection_status": "",
    "remark": "",
    "las_file_submitted": False,
    "reports_available_on_bms_for_las": False,
    "sar_files_submitted": False,
    "reports_available_on_bms_for_sar": False,
}


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _cell_str(value)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = _cell_str(value).lower()
    return s in ("yes", "y", "true", "1", "x")


def _normalize_inspection_status(raw: str) -> str:
    s = " ".join(raw.strip().split())
    if not s:
        return ""
    choices = [c[0] for c in AbstractStructureEntry.INSPECTION_STATUS_CHOICES]
    lower_map = {c.lower(): c for c in choices}
    return lower_map.get(s.lower(), s)


def _coerce(field: str, value: Any) -> Any:
    if field == "date_of_entry":
        return _parse_date(value)
    if field in (
        "las_file_submitted",
        "reports_available_on_bms_for_las",
        "sar_files_submitted",
        "reports_available_on_bms_for_sar",
    ):
        return _parse_bool(value)
    if field == "inspection_status":
        return _normalize_inspection_status(_cell_str(value))
    return _cell_str(value)


def _row_is_empty(payload: dict[str, Any]) -> bool:
    """Skip rows that contain no data at all (all optional fields may be blank)."""
    for key, value in payload.items():
        if key == "_excel_row":
            continue
        if isinstance(value, bool) and value:
            return False
        if value not in (None, "") and not (isinstance(value, str) and not value.strip()):
            return False
    return True


def _fallback_structure_type_from_third_column(data_row: tuple[Any, ...], col_to_field: dict[int, str]) -> str:
    """If no header mapped to structure_type, use Excel column C (3rd column) as Type."""
    if "structure_type" in col_to_field.values():
        return ""
    if len(data_row) < 3:
        return ""
    return _cell_str(data_row[2])


def parse_excel_workbook(file_bytes: bytes) -> tuple[list[dict[str, Any]], str | None]:
    """
    Returns (list of row dicts for serializer, error_message).
    Uses first worksheet; row 1 = headers; data from row 2.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface parse errors
        return [], f"Could not read Excel file: {exc}"

    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], "The sheet is empty."

        col_to_field: dict[int, str] = {}
        for idx, cell in enumerate(header_row):
            field = HEADER_TO_FIELD.get(_norm_header(cell))
            if field:
                col_to_field[idx] = field

        if not col_to_field:
            return [], "No recognized column headers in row 1. Use headers like Team lead name, Date, Chainage, …"

        out: list[dict[str, Any]] = []
        for excel_row_idx, data_row in enumerate(rows_iter, start=2):
            payload: dict[str, Any] = {}
            for col_idx, field in col_to_field.items():
                if col_idx < len(data_row):
                    payload[field] = _coerce(field, data_row[col_idx])
                else:
                    payload[field] = _coerce(field, None)

            # User format support: when headers are inconsistent, treat 3rd column as Type.
            fallback_type = _fallback_structure_type_from_third_column(data_row, col_to_field)
            if fallback_type and not str(payload.get("structure_type") or "").strip():
                payload["structure_type"] = fallback_type

            if _row_is_empty(payload):
                continue

            merged = {**ENTRY_PAYLOAD_DEFAULTS, **payload}
            if not (merged.get("branch") or "").strip():
                merged["branch"] = "INFRA_CORE"

            status = merged.get("inspection_status") or ""
            if status == AbstractStructureEntry.INSPECTION_COMPLETED:
                for tf in (
                    "las_file_submitted",
                    "reports_available_on_bms_for_las",
                    "sar_files_submitted",
                    "reports_available_on_bms_for_sar",
                ):
                    merged.setdefault(tf, False)

            out.append({"_excel_row": excel_row_idx, **merged})

        if not out:
            return [], "No data rows found after the header (or all rows were blank)."

        return out, None
    finally:
        wb.close()
